from __future__ import annotations

import hashlib
from dataclasses import dataclass
from openpyxl import load_workbook

from kbeton.importers.utils import norm_header, parse_date, parse_money

@dataclass
class FinanceRow:
    date: object
    amount: float
    currency: str
    description: str
    counterparty: str
    tx_type: str | None
    raw_fields: dict

FINANCE_HEADER_SYNONYMS = {
    "date": {"date", "дата", "дата документа", "период"},
    "amount": {"amount", "сумма", "сумма документа", "сумма операции", "сумма к оплате"},
    "currency": {"currency", "валюта", "вал"},
    "description": {"description", "назначение", "комментарий", "содержание", "операция", "основание"},
    "counterparty": {"counterparty", "контрагент", "наименование контрагента", "контрагент наименование"},
    "tx_type": {"type", "вид", "приход/расход", "движение", "тип"},
}

def bytes_to_filelike(data: bytes):
    from io import BytesIO
    return BytesIO(data)

def _find_header_row(rows: list[list[object]], max_scan: int = 15) -> tuple[int, dict[str, int]]:
    for i in range(min(max_scan, len(rows))):
        r = rows[i]
        headers = [norm_header(str(c)) if c is not None else "" for c in r]
        idx_map: dict[str, int] = {}
        for field, syns in FINANCE_HEADER_SYNONYMS.items():
            for col_i, h in enumerate(headers):
                if h in syns:
                    idx_map[field] = col_i
                    break
        if "amount" in idx_map and ("date" in idx_map or "description" in idx_map):
            return i, idx_map
    raise ValueError("Cannot detect header row for finance import")

def parse_finance_xlsx(data: bytes, *, default_currency: str = "KGS") -> list[FinanceRow]:
    wb = load_workbook(filename=bytes_to_filelike(data), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_idx, idx_map = _find_header_row(rows)
    out: list[FinanceRow] = []

    for r in rows[header_idx + 1 :]:
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        raw = {str(i): (c if c is not None else "") for i, c in enumerate(r)}
        dt = parse_date(r[idx_map['date']]) if 'date' in idx_map else None
        amt = parse_money(r[idx_map['amount']]) if 'amount' in idx_map else 0.0
        cur = str(r[idx_map['currency']]).strip() if 'currency' in idx_map and r[idx_map['currency']] else default_currency
        desc = str(r[idx_map['description']]).strip() if 'description' in idx_map and r[idx_map['description']] is not None else ''
        cp = str(r[idx_map['counterparty']]).strip() if 'counterparty' in idx_map and r[idx_map['counterparty']] is not None else ''
        ttype = str(r[idx_map['tx_type']]).strip().lower() if 'tx_type' in idx_map and r[idx_map['tx_type']] is not None else None
        out.append(FinanceRow(date=dt, amount=amt, currency=cur or default_currency, description=desc, counterparty=cp, tx_type=ttype, raw_fields=raw))
    return out

def make_dedup_hash(row: FinanceRow) -> str:
    base = f"{row.date}|{row.amount}|{row.currency}|{row.description}|{row.counterparty}|{row.tx_type}"
    return hashlib.sha256(base.encode("utf-8")).hexdigest()
