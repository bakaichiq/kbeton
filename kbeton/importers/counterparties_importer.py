from __future__ import annotations

from dataclasses import dataclass
from openpyxl import load_workbook

from kbeton.importers.utils import norm_header, parse_date, parse_money, norm_counterparty_name

@dataclass
class CounterpartyRow:
    counterparty_name: str
    counterparty_name_norm: str
    receivable_money: float
    receivable_assets: str
    payable_money: float
    payable_assets: str
    ending_balance_money: float
    raw_fields: dict

CP_HEADER_SYNONYMS = {
    "counterparty_name": {"counterparty_name", "контрагент", "наименование контрагента", "контрагент наименование", "наименование"},
    "receivable_money": {"receivable_money", "дебиторка", "нам должны деньги", "задолженность нам (деньги)", "дебиторская задолженность (деньги)"},
    "receivable_assets": {"receivable_assets", "нам должны активами", "дебиторка (активы)", "задолженность нам (активы)"},
    "payable_money": {"payable_money", "кредиторка", "мы должны деньги", "задолженность наша (деньги)", "кредиторская задолженность (деньги)"},
    "payable_assets": {"payable_assets", "мы должны активами", "кредиторка (активы)", "задолженность наша (активы)"},
    "ending_balance_money": {"ending_balance_money", "сальдо конечное", "конечное сальдо", "итого", "сальдо"},
}

def bytes_to_filelike(data: bytes):
    from io import BytesIO
    return BytesIO(data)

def _find_header_row(rows: list[list[object]], max_scan: int = 20) -> tuple[int, dict[str, int]]:
    for i in range(min(max_scan, len(rows))):
        r = rows[i]
        headers = [norm_header(str(c)) if c is not None else "" for c in r]
        idx_map: dict[str, int] = {}
        for field, syns in CP_HEADER_SYNONYMS.items():
            for col_i, h in enumerate(headers):
                if h in syns:
                    idx_map[field] = col_i
                    break
        if "counterparty_name" in idx_map and ("ending_balance_money" in idx_map or "receivable_money" in idx_map or "payable_money" in idx_map):
            return i, idx_map
    raise ValueError("Cannot detect header row for counterparties import")

def parse_counterparties_xlsx(data: bytes) -> list[CounterpartyRow]:
    wb = load_workbook(filename=bytes_to_filelike(data), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_idx, idx_map = _find_header_row(rows)
    out: list[CounterpartyRow] = []

    for r in rows[header_idx + 1 :]:
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        raw = {str(i): (c if c is not None else "") for i, c in enumerate(r)}
        name = str(r[idx_map["counterparty_name"]]).strip() if r[idx_map["counterparty_name"]] is not None else ""
        out.append(
            CounterpartyRow(
                counterparty_name=name,
                counterparty_name_norm=norm_counterparty_name(name),
                receivable_money=parse_money(r[idx_map["receivable_money"]]) if "receivable_money" in idx_map else 0.0,
                receivable_assets=str(r[idx_map["receivable_assets"]]).strip() if "receivable_assets" in idx_map and r[idx_map["receivable_assets"]] is not None else "",
                payable_money=parse_money(r[idx_map["payable_money"]]) if "payable_money" in idx_map else 0.0,
                payable_assets=str(r[idx_map["payable_assets"]]).strip() if "payable_assets" in idx_map and r[idx_map["payable_assets"]] is not None else "",
                ending_balance_money=parse_money(r[idx_map["ending_balance_money"]]) if "ending_balance_money" in idx_map else 0.0,
                raw_fields=raw,
            )
        )
    return out
