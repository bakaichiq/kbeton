from __future__ import annotations

from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from kbeton.reports.pnl import PnlRow

def pnl_to_xlsx(rows: list[PnlRow], *, period: str, start, end, totals: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"

    ws.append(["Period", "Income", "Expense", "Net"])
    for r in rows:
        ws.append([r.period_start.isoformat(), round(r.income_sum, 2), round(r.expense_sum, 2), round(r.net_profit, 2)])

    ws.append([])
    ws.append(["TOTAL", round(totals["total_income"], 2), round(totals["total_expense"], 2), round(totals["total_net"], 2)])
    ws.append(["UNKNOWN_ROWS", totals.get("unknown_count", 0), "", ""])

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Daily dynamics sheet
    daily = totals.get("daily", [])
    ws_daily = wb.create_sheet("Daily")
    ws_daily.append(["Date", "Income", "Expense", "Net"])
    for r in daily:
        ws_daily.append([r["date"].isoformat(), round(r["income"], 2), round(r["expense"], 2), round(r["net"], 2)])
    for col in range(1, 5):
        ws_daily.column_dimensions[get_column_letter(col)].width = 18

    # Top articles sheet
    ws_top = wb.create_sheet("Top Articles")
    ws_top.append(["Type", "Article", "Amount"])
    for r in totals.get("top_income_articles", []):
        ws_top.append(["income", r["name"], round(r["amount"], 2)])
    for r in totals.get("top_expense_articles", []):
        ws_top.append(["expense", r["name"], round(r["amount"], 2)])
    for col in range(1, 4):
        ws_top.column_dimensions[get_column_letter(col)].width = 22

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
