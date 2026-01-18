from __future__ import annotations

from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def production_shifts_to_xlsx(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Shifts"
    ws.append(["Shift ID", "Date", "Shift", "Line", "Operator", "Product", "Mark", "Qty", "UOM"])
    for r in rows:
        ws.append([
            r.get("shift_id"),
            r.get("date"),
            r.get("shift_type"),
            r.get("line"),
            r.get("operator"),
            r.get("product"),
            r.get("mark"),
            round(float(r.get("qty", 0)), 3),
            r.get("uom"),
        ])
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
